import openpyxl as op
import xlrd
excel_data_file = xlrd.open_workbook('marketdo4a-spb-riverstart.xlsx')
sheet = excel_data_file.sheet_by_index(0)
row_numbers = sheet.nrows
print(row_numbers)

wb = op.load_workbook('marketdo4a-spb-riverstart.xlsx')
ws = wb['Sheet1']

#переводим заголовки
ws.cell(row = 1, column = 1).value = 'Кампания'
ws.cell(row = 1, column = 2).value = 'Показы'
ws.cell(row = 1, column = 3).value = 'Переходы/Клики'
ws.cell(row = 1, column = 4).value = 'CTR'
ws.cell(row = 1, column = 5).value = 'Бюджет'
ws.cell(row = 1, column = 6).value = 'Продажи с сайта'
ws.cell(row = 1, column = 7).value = 'Доход с сайта'
ws.cell(row = 1, column = 8).value = 'CPC'
ws.cell(row = 1, column = 9).value = 'CPO'
ws.cell(row = 1, column = 10).value = 'CV'

true_row_number = row_numbers - 2

#информация о логике циклов дана в цикле расчет сро
#Цикл расчета cpc (считает python)
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
cpc = ws['H2']
cpo = ws['I2']
cv = ws['J2']

a = 1
for i in range(true_row_number):
    a += 1
    clicks = ws['C' + str(a)]
    budget_col = ws['E' + str(a)]
    val_if_zero = 0
    if clicks.value == 0:
        ws.cell(row = a, column = 8).value = val_if_zero
    else:
        result_cpc = budget_col.value / clicks.value
        ws.cell(row = a, column = 8).value = result_cpc
    pass

#Цикл расчета СPO (считает python)
campaign_name = ws['A2'] #объявляем название показателей как переменных (A2 потому что не берем строку с заголовками столбцов)
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
cpc = ws['H2']
cpo = ws['I2']
cv = ws['J2']

a = 1 #переменная 'a' - динамическая, используется для отсчета номера ряда в экселе. 
for i in range(true_row_number):
    a += 1 #на каждой итерации цикла прибавляем единицу к динамической переменной чтобы записывать данные в нужный ряд
    budget_col = ws['E' + str(a)] #в данном случае считаем cpo, поэтому берем два показателя, и указываем для них координаты. Причем строка - это динамическая переменная а.
    ecomm_sales = ws['F' + str(a)]
    val_if_zero = 0
    if ecomm_sales.value == '--' or ecomm_sales.value == 0: #если нет данных, или стоит 0, будем в результат тоже записывать 0
        ws.cell(row = a, column = 6).value = val_if_zero
        ws.cell(row = a, column = 9).value = val_if_zero
    else:
        result_cpo = budget_col.value / ecomm_sales.value #в остальных случаях расчет возможен, поэтому пишем формулу расчета cpo(при чем делимое и делитель постоянно меняются за счет динамической переменной a)
        ws.cell(row = a, column = 9).value = result_cpo #результат равен частному из предыдущей строки, записываем его в ячейку таблицы.
    pass

#Цикл расчета CV (считает python)
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
cpc = ws['H2']
cpo = ws['I2']
cv = ws['J2']

a = 1
for i in range(true_row_number):
    a += 1
    clicks = ws['C' + str(a)]
    ecomm_sales = ws['F' + str(a)]
    val_if_zero = 0
    if clicks.value == '--' or clicks.value == 0:
        ws.cell(row = a, column = 10).value = val_if_zero
    else:
        result_cv = ecomm_sales.value / clicks.value
        ws.cell(row = a, column = 10).value = result_cv
    pass

#приводим к нулю показатели дохода если дохода нет
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
cpc = ws['H2']
cpo = ws['I2']
cv = ws['J2']

a = 1
for i in range(true_row_number):
    a += 1
    money = ws['G' + str(a)]
    val_if_zero = 0
    if money.value == '--':
        ws.cell(row = a, column = 7).value = val_if_zero
    else:
        pass
    pass
    
wb.save('marketdo4a-spb-riverstart.xlsx')
wb.close()