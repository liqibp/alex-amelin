# -*- coding: utf-8 -*-
import requests
from requests.exceptions import ConnectionError
from time import sleep
import json
import datetime
from xlrd import open_workbook
import xlsxwriter
import csv
import openpyxl as op
import xlrd
import os
import random

# Метод для корректной обработки строк в кодировке UTF-8 как в Python 3, так и в Python 2
import sys

if sys.version_info < (3,):
    def u(x):
        try:
            return x.encode("utf8")
        except UnicodeDecodeError:
            return x
else:
    def u(x):
        if type(x) == type(b''):
            return x.decode('utf8')
        else:
            return x

# Получаем вчерашнюю дату
def getYesterday(): 
    today=datetime.date.today() 
    oneday=datetime.timedelta(days=1) 
    yesterday=today-oneday  
    return yesterday
 
date1 = str(getYesterday())

# --- Входные данные ---
# Адрес сервиса Reports для отправки JSON-запросов (регистрозависимый)
ReportsURL = 'https://api.direct.yandex.com/json/v5/reports'

# OAuth-токен пользователя, от имени которого будут выполняться запросы
token = 'AQAAAAAye9OuAAWoSe5kDEDFiE9IkbFVe-zMvf0'

# Логин клиента рекламного агентства
# Обязательный параметр, если запросы выполняются от имени рекламного агентства
clientLogin = str(input("Введите логин клиента: "))

# --- Подготовка запроса ---
# Создание HTTP-заголовков запроса
headers = {
           # OAuth-токен. Использование слова Bearer обязательно
           "Authorization": "Bearer " + token,
           # Логин клиента рекламного агентства
           "Client-Login": clientLogin,
           # Язык ответных сообщений
           "Accept-Language": "ru",
           # Режим формирования отчета
           "processingMode": "auto",
           # Формат денежных значений в отчете
           "returnMoneyInMicros": "false",
           # Не выводить в отчете строку с названием отчета и диапазоном дат
           "skipReportHeader": "true"
           # Не выводить в отчете строку с названиями полей
           # "skipColumnHeader": "true"
           # Не выводить в отчете строку с количеством строк статистики
           # "skipReportSummary": "true"
           }

#Создаем переменную, строковое значение которой случайно генерируется. Дает возможность постоянно менять название отчета, чтобы сервер яндекса заново обрабатывал каждый раз
some_number = random.randrange(1, 999, 1)
report_name = str(clientLogin) + str(some_number)

#Запрашиваем дату у пользователя
while True:
    print("Выберите временной диапазон")
    choose_date = str(input("1 - Статистика за вчерашний день, 2 - Свой диапазон: "))
    if choose_date == '1':
        date_user1 = date1
        date_user2 = date1
        print("Okay, отправляю запрос на сервер яндекса...")
        break
    elif choose_date == '2':
        print('Введите начало и конец периода. Даты задаются в формате DD-MM-YYYY')
        dd1 = int(input("Введите начальную дату(день): "))
        if dd1 > 31:
            while dd1 > 31:
                print('Неверный день. только 1 - 31')
                dd1 = int(input("Введите начальную дату(день): "))
            else:
                pass
        else:
            pass

        day1 = dd1
        day1 = str(day1)
        if dd1 < 10:
            day1 = "0" + day1
        else:
            pass

        mm1 = int(input("Введите начальную дату(месяц): "))
        if dd1 == 29 or dd1 == 30 or dd1 == 31:
            while mm1 == 2:
                print('В феврале нет столько дней. введите другое значение')
                mm1 = int(input("Введите начальную дату(месяц): "))
            else:
                pass
        elif dd1 == 31:
            while mm1 == 4 or mm1 == 6 or mm1 == 9 or mm1 == 11:
                print('В месяце нет столько дней. введите другое значение')
                mm1 = int(input("Введите начальную дату(месяц): "))
            else:
                pass
        elif mm1 > 12:
            while mm1 > 12:
                print('Неверный месяц. введите заново')
                mm1 = int(input("Введите начальную дату(месяц): "))
            else:
                pass
        else:
            pass

        month1 = mm1
        month1 = str(month1)
        if mm1 < 10:
            month1 = "0" + month1
        else:
            pass

        yyyy1 = int(input("Введите начальную дату(год): "))

        print("okay")

        dd2 = int(input("Введите конечную дату(день): "))
        if dd2 > 31:
            while dd1 > 31:
                print('Неверный день. только 1 - 31')
                dd2 = int(input("Введите конечную дату(день): "))
            else:
                pass
        else:
            pass

        day2 = dd2
        day2 = str(day2)
        if dd2 < 10:
            day2 = "0" + day2
        else:
            pass

        mm2 = int(input("Введите конечную дату(месяц): "))
        if dd2 == 29 or dd2 == 30 or dd2 == 31:
            while mm2 == 2:
                print('В феврале нет столько дней. введите другое значение')
                mm2 = int(input("Введите конечную дату(месяц): "))
            else:
                pass
        elif dd2 == 31:
            while mm2 == 4 or mm2 == 6 or mm2 == 9 or mm2 == 11:
                print('В месяце нет столько дней. введите другое значение')
                mm2 = int(input("Введите конечную дату(месяц): "))
            else:
                pass
        elif mm2 > 12:
            while mm2 > 12:
                print('Неверный месяц. введите заново')
                mm2 = int(input("Введите конечную дату(месяц): "))
            else:
                pass
        else:
            pass

        month2 = mm2
        month2 = str(month2)
        if mm2 < 10:
            month2 = "0" + month2
        else:
            pass

        yyyy2 = int(input("Введите конечную дату(год): "))
        if dd2 < dd1 and mm2 <= mm1 and yyyy2 <= yyyy1:
            while dd2 < dd1 and mm2 <= mm1 and yyyy2 <= yyyy1:
                print('Год не может быть меньше при такой дате. Введите другой год')
                yyyy2 = int(input("Введите конечную дату(год): "))
            else:
                pass
        else:
            pass

        str(yyyy1)
        str(yyyy2)

        date_user1 = (str(yyyy1) + "-" + str(month1) + "-" + str(day1))
        date_user2 = (str(yyyy2) + "-" + str(month2) + "-" + str(day2))
        break
    else:
        print("Неверный выбор. Введите число - 1 или 2")

#Запрашиваем номер цели 
goal_number = str(input("Введите номер цели, для получения данных. Если цель не нужна, поставьте - : "))
if goal_number == "-":
    body = {
        "params": {
            "SelectionCriteria": {
                "DateFrom": date_user1,
                "DateTo": date_user2
            },
            "FieldNames": [
                "CampaignName",
                "Impressions",
                "Clicks",
                "Ctr",
                "Cost",
                "Conversions",
                "Revenue",
                "BounceRate"
            ],
            "ReportName": u(report_name),
            "ReportType": "CAMPAIGN_PERFORMANCE_REPORT",
            "DateRangeType": "CUSTOM_DATE",
            "Format": "TSV",
            "IncludeVAT": "YES",
            "IncludeDiscount": "NO"
        }
    }
else:
    body = {
        "params": {
            "SelectionCriteria": {
                "DateFrom": date_user1,
                "DateTo": date_user2
            },
            "Goals": [
                goal_number #3024824726 - для дочи
            ],
            "FieldNames": [
                "CampaignName",
                "Impressions",
                "Clicks",
                "Ctr",
                "Cost",
                "Conversions",
                "Revenue",
                "BounceRate"
            ],
            "ReportName": u(report_name),
            "ReportType": "CAMPAIGN_PERFORMANCE_REPORT",
            "DateRangeType": "CUSTOM_DATE",
            "Format": "TSV",
            "IncludeVAT": "YES",
            "IncludeDiscount": "NO"
        }
    }
    pass

print("okay, отправляю запрос на сервер яндекса...")
# Кодирование тела запроса в JSON
body = json.dumps(body, indent=4)

# --- Запуск цикла для выполнения запросов ---
# Если получен HTTP-код 200, то выводится содержание отчета
# Если получен HTTP-код 201 или 202, выполняются повторные запросы
while True:
    try:
        req = requests.post(ReportsURL, body, headers=headers)
        req.encoding = 'utf-8'  # Принудительная обработка ответа в кодировке UTF-8
        if req.status_code == 400:
            print("Параметры запроса указаны неверно или достигнут лимит отчетов в очереди")
            print("RequestId: {}".format(req.headers.get("RequestId", False)))
            print("JSON-код запроса: {}".format(u(body)))
            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
            break
        elif req.status_code == 200:
            format(u(req.text))
            break
        elif req.status_code == 201:
            print("Отчет успешно поставлен в очередь в режиме офлайн")
            retryIn = int(req.headers.get("retryIn", 60))
            print("Повторная отправка запроса через {} секунд".format(retryIn))
            print("RequestId: {}".format(req.headers.get("RequestId", False)))
            sleep(retryIn)
        elif req.status_code == 202:
            print("Отчет формируется в режиме офлайн")
            retryIn = int(req.headers.get("retryIn", 60))
            print("Повторная отправка запроса через {} секунд".format(retryIn))
            print("RequestId:  {}".format(req.headers.get("RequestId", False)))
            sleep(retryIn)
        elif req.status_code == 500:
            print("При формировании отчета произошла ошибка. Пожалуйста, попробуйте повторить запрос позднее")
            print("RequestId: {}".format(req.headers.get("RequestId", False)))
            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
            break
        elif req.status_code == 502:
            print("Время формирования отчета превысило серверное ограничение.")
            print("Пожалуйста, попробуйте изменить параметры запроса - уменьшить период и количество запрашиваемых данных.")
            print("JSON-код запроса: {}".format(body))
            print("RequestId: {}".format(req.headers.get("RequestId", False)))
            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
            break
        else:
            print("Произошла непредвиденная ошибка")
            print("RequestId:  {}".format(req.headers.get("RequestId", False)))
            print("JSON-код запроса: {}".format(body))
            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
            break

    # Обработка ошибки, если не удалось соединиться с сервером API Директа
    except ConnectionError:
        # В данном случае мы рекомендуем повторить запрос позднее
        print("Произошла ошибка соединения с сервером API")
        # Принудительный выход из цикла
        break

    # Если возникла какая-либо другая ошибка
    except:
        # В данном случае мы рекомендуем проанализировать действия приложения
        print("Произошла непредвиденная ошибка")
        # Принудительный выход из цикла
        break

#Задаем параметры книги xls
# file = open("marketdo4a-chlb-riverstart.csv", "w")
# string = req.text
# string = string.replace(",", ".")
# string = string.replace("\t", ",")
# print(string)
# file.write(string)
# file.close()
file = open(str(clientLogin) + '.tsv', "w")

#в полученных данных меняем символ точки на запятую, для корректной работы в excel
string = req.text
string = string.replace(".", ",")
file.write(string)
file.close()

print("Строю таблицу в эксельке...")

#Запись tsv данных в xlsx
from xlsxwriter.workbook import Workbook
# Add some command-line logic to read the file names.
tsv_file = str(clientLogin) + '.tsv'
xlsx_file = str(clientLogin) + '.xlsx'

# Create an XlsxWriter workbook object and add a worksheet.
workbook = Workbook(xlsx_file)
worksheet = workbook.add_worksheet()

# Create a TSV file reader.
tsv_reader = csv.reader(open(tsv_file, 'rt'), delimiter='\t')

# Read the row data from the TSV file and write it to the XLSX file.
for row, data in enumerate(tsv_reader):
    worksheet.write_row(row, 0, data)

# Close the XLSX file.
workbook.close()

#Обрабатываем полученный xlsx файл
excel_data_file = xlrd.open_workbook(str(clientLogin) + '.xlsx')
sheet = excel_data_file.sheet_by_index(0)
row_numbers = sheet.nrows
print("Всего строк в таблице: " + str(row_numbers))

wb = op.load_workbook(str(clientLogin) + '.xlsx')
ws = wb['Sheet1']

#переводим заголовки
ws.cell(row = 1, column = 1).value = 'Кампания'
ws.cell(row = 1, column = 2).value = 'Показы'
ws.cell(row = 1, column = 3).value = 'Переходы/Клики'
ws.cell(row = 1, column = 4).value = 'CTR'
ws.cell(row = 1, column = 5).value = 'Бюджет'
ws.cell(row = 1, column = 6).value = 'Конверсии'
ws.cell(row = 1, column = 7).value = 'Доход'
ws.cell(row = 1, column = 8).value = '% Отказов'
ws.cell(row = 1, column = 9).value = 'CPC'
ws.cell(row = 1, column = 10).value = 'CPO'
ws.cell(row = 1, column = 11).value = 'CV'

true_row_number = row_numbers - 2

print("Считаю показатели CPC, CPO, CV...")
#информация о логике циклов дана в цикле расчет сро
#Цикл расчета cpc (считает excel)
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
bounce_rate = ws['H2']
cpc = ws['I2']
cpo = ws['J2']
cv = ws['K2']

a = 1
for i in range(true_row_number):
    a += 1
    clicks = ws['C' + str(a)]
    budget_col = ws['E' + str(a)]
    val_if_zero = 0
    if clicks.value == '--' or clicks.value == 0 or clicks.value == '0':
        ws.cell(row = a, column = 9).value = val_if_zero
    else:
        b = '=E' + str(a) + '/C' + str(a)
        ws.cell(row = a, column = 9).value = str(b)
    pass

#Цикл расчета СPO (считает excel)
campaign_name = ws['A2'] #объявляем название показателей как переменных (A2 потому что не берем строку с заголовками столбцов)
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
bounce_rate = ws['H2']
cpc = ws['I2']
cpo = ws['J2']
cv = ws['K2']

a = 1 #переменная 'a' - динамическая, используется для отсчета номера ряда в экселе. 
for i in range(true_row_number):
    a += 1 #на каждой итерации цикла прибавляем единицу к динамической переменной чтобы записывать данные в нужный ряд
    budget_col = ws['E' + str(a)] #в данном случае считаем cpo, поэтому берем два показателя, и указываем для них координаты. Причем строка - это динамическая переменная а.
    ecomm_sales = ws['F' + str(a)]
    val_if_zero = 0
    if ecomm_sales.value == '--' or ecomm_sales.value == 0 or ecomm_sales.value == '0': #если нет данных, или стоит 0, будем в результат тоже записывать 0
        ws.cell(row = a, column = 6).value = val_if_zero
        ws.cell(row = a, column = 10).value = val_if_zero
    else:
        b = '=E' + str(a) + '/F' + str(a)
        ws.cell(row = a, column = 10).value = str(b) #результат равен частному из предыдущей строки, записываем его в ячейку таблицы.
    pass

#Цикл расчета CV (считает excel)
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
bounce_rate = ws['H2']
cpc = ws['I2']
cpo = ws['J2']
cv = ws['K2']

a = 1
for i in range(true_row_number):
    a += 1
    clicks = ws['C' + str(a)]
    ecomm_sales = ws['F' + str(a)]
    val_if_zero = 0
    if clicks.value == '--' or clicks.value == 0 or clicks.value == '0':
        ws.cell(row = a, column = 11).value = val_if_zero
    else:
        b = '=F' + str(a) + '/C' + str(a)
        ws.cell(row = a, column = 11).value = str(b)
    pass

#приводим к нулю показатели дохода если дохода нет
campaign_name = ws['A2']
impressions = ws['B2']
clicks = ws['C2']
ctr_val = ws['D2']
budget_col = ws['E2']
ecomm_sales = ws['F2']
money = ws['G2']
bounce_rate = ws['H2']
cpc = ws['I2']
cpo = ws['J2']
cv = ws['K2']

a = 1
for i in range(true_row_number):
    a += 1
    money = ws['G' + str(a)]
    val_if_zero = 0
    if money.value == '--' or money.value == 0 or money.value == '0':
        ws.cell(row = a, column = 7).value = val_if_zero
    else:
        pass
    pass

wb.save(str(clientLogin) + '.xlsx')
wb.close()

print("Отчет готов. Откройте xlsx файл с названием логина клиента")
input("Нажмите Enter для выхода ")